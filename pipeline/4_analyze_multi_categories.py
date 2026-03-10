import pandas as pd
import numpy as np
import sys
import io
import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList

from pathlib import Path
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# Set up paths
script_dir = os.path.dirname(os.path.abspath(__file__))
base_dir = os.path.dirname(script_dir)  # Go up two levels from 
data_file = os.path.join(base_dir, 'processed_data', 'full_data_multi_category.xlsx')
output_dir = os.path.join(base_dir, 'output')
output_file = os.path.join(output_dir, 'MULTI_CATEGORY_ANALYSIS.xlsx')

# Create output directory if it doesn't exist
os.makedirs(output_dir, exist_ok=True)

print("="*80)
print("CREATING MULTI-CATEGORY ANALYSIS")
print("="*80)
print("Goal: One sheet per question with general + demographic analysis")
print("Feature: Multi-category responses counted separately")
print("Source: full_data_multi_category.xlsx (Multi-category LLM)\n")

# Load data
print(f"Loading data from: {data_file}")
df = pd.read_excel(data_file)

# First row is metadata/headers, rest is data
metadata = df.iloc[0]
responses = df.iloc[1:].reset_index(drop=True)

print(f"Total respondents: {len(responses)}\n")

# ============================================================
# PROCESS DEMOGRAPHICS
# ============================================================
print("Processing demographics...")

# NPS Tier (col 4)
nps_tier = responses.iloc[:, 4].fillna('Missing')

# Gender (cols 44-47) - multi-select - INCLUDE ALL OPTIONS
gender_cols = [44, 45, 46, 47]  # Female, Male, Other, Prefer not to say
gender_options = [metadata.iloc[i] for i in gender_cols]

gender_selections = []
for idx in range(len(responses)):
    selected = []
    for i, col_idx in enumerate(gender_cols):
        if pd.notna(responses.iloc[idx, col_idx]):
            selected.append(gender_options[i])
    if len(selected) == 0:
        primary = "Not specified"
    elif len(selected) == 1:
        primary = selected[0]
    else:
        primary = "Multiple"
    gender_selections.append(primary)

primary_gender = pd.Series(gender_selections)

# Age (col 43)
age_group = responses.iloc[:, 43].fillna('Not specified')

# Payer Segment (col 51) - THE REAL DATA!
payer_segment = responses.iloc[:, 51].fillna('Unknown')

# Add to dataframe
responses['NPS_Tier'] = nps_tier
responses['Primary_Gender'] = primary_gender
responses['Age_Group'] = age_group
responses['Payer_Segment'] = payer_segment

print("✓ Demographics processed\n")

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def add_section(rows, title, is_header=True):
    """Add a formatted section header"""
    rows.append([''])
    if is_header:
        rows.append(['═'*100])
        rows.append([title])
        rows.append(['═'*100])
    else:
        rows.append([title])
    return rows

def split_multi_categories(insight_series):
    """
    Split multi-category strings and return Series with all individual category mentions.
    Input: Series with values like "Category1 | Category2"
    Output: Series with individual categories (longer than input)
    """
    all_categories = []
    for val in insight_series:
        if pd.notna(val):
            # Split by delimiter
            cats = [c.strip() for c in str(val).split('|')]
            all_categories.extend(cats)
    return pd.Series(all_categories)

def create_respondent_summary_sheet():
    """Create summary sheet showing respondent counts for all questions"""
    rows = []

    rows.append(['SURVEY RESPONSE SUMMARY'])
    rows.append([''])
    rows.append([f'Total Survey Respondents: {len(responses)}'])
    rows.append([''])

    add_section(rows, 'RESPONDENT COUNTS BY QUESTION', True)

    # Define questions
    questions = [
        (3, "Q2", "NPS Score (0-10)", "rating"),
        (5, "Q3", "What did you like?", "open-ended"),
        (7, "Q4", "How can we improve?", "open-ended"),
        (9, "Q5", "Technical Satisfaction (1-5)", "rating"),
        (11, "Q6", "Fairness Satisfaction (1-5)", "rating"),
        (13, "Q7", "Progression Satisfaction (1-5)", "rating"),
        (list(range(15, 29)), "Q8", "Play Motivations", "multiselect"),
        (list(range(29, 39)), "Q9", "Favorite Modes", "multiselect"),
        (39, "Q10", "Game Mode Variety", "categorical"),
        (40, "Q11", "Update Frequency", "categorical"),
        (41, "Q12", "One thing to change", "open-ended"),
        (43, "Q13", "Age", "demographic"),
        (list(range(44, 48)), "Q14", "Gender", "demographic-multi"),
        (48, "Q15", "Final comments", "open-ended"),
    ]

    rows.append(['Question', 'Answered', 'Skipped', '% Skipped'])

    for col_info, q_num, q_text, q_type in questions:
        if q_type == "multiselect" or q_type == "demographic-multi":
            answered = responses.iloc[:, col_info].notna().any(axis=1).sum()
        else:
            answered = responses.iloc[:, col_info].notna().sum()

        skipped = len(responses) - answered
        skip_pct = skipped / len(responses)

        rows.append([f'{q_num}: {q_text}', answered, skipped, skip_pct])

    rows.append([''])

    # Add demographic distributions
    add_section(rows, 'DEMOGRAPHIC DISTRIBUTIONS', True)

    # Gender
    rows.append(['GENDER DISTRIBUTION'])
    rows.append(['Gender', 'Count', '% of Total'])
    gender_dist = responses['Primary_Gender'].value_counts()
    for gender, count in gender_dist.items():
        pct = count / len(responses)
        rows.append([gender, count, pct])
    rows.append(['TOTAL', len(responses), 1.0])
    rows.append([''])

    # Age
    rows.append(['AGE DISTRIBUTION'])
    rows.append(['Age Group', 'Count', '% of Total'])
    age_dist = responses['Age_Group'].value_counts()
    for age, count in age_dist.items():
        pct = count / len(responses)
        rows.append([age, count, pct])
    rows.append(['TOTAL', len(responses), 1.0])
    rows.append([''])

    # Payer Segment
    rows.append(['PAYER SEGMENT DISTRIBUTION'])
    rows.append(['Segment', 'Count', '% of Total'])
    payer_dist = responses['Payer_Segment'].value_counts()
    for segment, count in payer_dist.items():
        pct = count / len(responses)
        rows.append([segment, count, pct])
    rows.append(['TOTAL', len(responses), 1.0])

    return rows


def create_openended_sheet(q_num, question_text, answer_col, insight_col):
    """Create sheet for open-ended questions with themes (MULTI-CATEGORY SUPPORT)"""
    rows = []

    # Title
    rows.append([f'Q#{q_num}: {question_text}'])
    rows.append([''])

    # PART 1: GENERAL ANALYSIS
    add_section(rows, 'GENERAL ANALYSIS - Overall Theme Distribution (Multi-Category)', True)

    total_respondents = len(responses)
    total_answers = responses.iloc[:, answer_col].notna().sum()
    no_response = total_respondents - total_answers
    no_response_pct = (no_response / total_respondents * 100) if total_respondents > 0 else 0

    # Get insights for people who answered
    answered_insights = responses[responses.iloc[:, answer_col].notna()].iloc[:, insight_col]

    # Split multi-categories and count each mention
    all_category_mentions = split_multi_categories(answered_insights)
    theme_counts = all_category_mentions.value_counts()  # All themes, sorted by count descending

    # Calculate substantive responses
    trivial_categories = ['No Additional Feedback', 'No Response']
    # Count how many responses had ONLY trivial categories
    trivial_only_count = 0
    for val in answered_insights:
        if pd.notna(val):
            cats = [c.strip() for c in str(val).split('|')]
            if all(c in trivial_categories for c in cats):
                trivial_only_count += 1

    substantive_answers = total_answers - trivial_only_count

    rows.append(['Theme', 'Mentions', '% of All Answers', '% of Substantive'])
    for theme, count in theme_counts.items():
        if pd.notna(theme):
            # Store as decimal for Excel percentage formatting
            pct_all = count / total_answers if total_answers > 0 else 0
            # For trivial categories, don't show substantive %
            if theme in trivial_categories:
                rows.append([theme, count, pct_all, '-'])
            else:
                # CORRECTED: Percentage of substantive RESPONDENTS (not mentions)
                # This can exceed 100% total because one respondent can have multiple categories
                pct_substantive = count / substantive_answers if substantive_answers > 0 else 0
                rows.append([theme, count, pct_all, pct_substantive])

    rows.append([''])
    rows.append([f'Total respondents: {total_respondents}'])
    rows.append([f'Total answers provided: {total_answers}'])
    rows.append([f'  - Substantive answers: {substantive_answers}'])
    rows.append([f'  - Trivial responses: {trivial_only_count}'])
    rows.append([f'Total category mentions: {len(all_category_mentions)}'])
    rows.append([f'No Response (did not answer): {no_response} ({no_response_pct:.1f}%)'])
    rows.append([''])
    rows.append(['Note: A single response can have multiple categories'])
    rows.append(['Mentions column shows total times each category was mentioned'])

    # PART 2: DEMOGRAPHIC BREAKDOWNS
    add_section(rows, 'DEMOGRAPHIC ANALYSIS', True)

    # Create table with all demographics - SHOW ALL THEMES (not just top 10)
    all_themes = theme_counts.index.tolist()

    # By Gender - INCLUDE ALL OPTIONS
    add_section(rows, 'By Gender - % of substantive respondents by each gender', False)
    header = ['Theme']
    genders = ['Female', 'Male', 'Other', 'Prefer not to say', 'Multiple', 'Not specified']
    # Get respondents who answered this question for filtering
    respondents_who_answered = responses[responses.iloc[:, answer_col].notna()]

    for g in genders:
        # Count only those of this gender who answered this question
        gender_answered = respondents_who_answered[respondents_who_answered['Primary_Gender'] == g]
        # Count substantive RESPONDENTS (not mentions)
        substantive_count = 0
        for val in gender_answered.iloc[:, insight_col]:
            if pd.notna(val):
                cats = [c.strip() for c in str(val).split('|')]
                # Has at least one non-trivial category
                if any(c not in trivial_categories for c in cats):
                    substantive_count += 1
        if substantive_count > 0:
            header.append(f'{g}\n(n={substantive_count})')
    rows.append(header)

    for theme in all_themes:
        if pd.isna(theme):
            continue
        row = [theme]
        for g in genders:
            gender_answered = respondents_who_answered[respondents_who_answered['Primary_Gender'] == g]
            # Count substantive respondents
            substantive_count = 0
            for val in gender_answered.iloc[:, insight_col]:
                if pd.notna(val):
                    cats = [c.strip() for c in str(val).split('|')]
                    if any(c not in trivial_categories for c in cats):
                        substantive_count += 1
            if substantive_count > 0:
                # Count RESPONDENTS who have this theme (not mentions)
                count = 0
                for val in gender_answered.iloc[:, insight_col]:
                    if pd.notna(val):
                        cats = [c.strip() for c in str(val).split('|')]
                        if theme in cats:
                            count += 1
                pct = count / substantive_count  # Store as decimal for Excel formatting
                row.append(pct)
        rows.append(row)

    rows.append([''])

    # By Payer Segment
    add_section(rows, 'By Payer Segment - % of substantive respondents by each segment', False)
    header = ['Theme']
    payer_segs = ['Non_Payer', 'Low', 'Med', 'High', 'VIP']
    for ps in payer_segs:
        ps_answered = respondents_who_answered[respondents_who_answered['Payer_Segment'] == ps]
        # Count substantive RESPONDENTS (not mentions)
        substantive_count = 0
        for val in ps_answered.iloc[:, insight_col]:
            if pd.notna(val):
                cats = [c.strip() for c in str(val).split('|')]
                if any(c not in trivial_categories for c in cats):
                    substantive_count += 1
        header.append(f'{ps}\n(n={substantive_count})')
    rows.append(header)

    for theme in all_themes:
        if pd.isna(theme):
            continue
        row = [theme]
        for ps in payer_segs:
            ps_answered = respondents_who_answered[respondents_who_answered['Payer_Segment'] == ps]
            # Count substantive respondents
            substantive_count = 0
            for val in ps_answered.iloc[:, insight_col]:
                if pd.notna(val):
                    cats = [c.strip() for c in str(val).split('|')]
                    if any(c not in trivial_categories for c in cats):
                        substantive_count += 1
            if substantive_count > 0:
                # Count RESPONDENTS who have this theme
                count = 0
                for val in ps_answered.iloc[:, insight_col]:
                    if pd.notna(val):
                        cats = [c.strip() for c in str(val).split('|')]
                        if theme in cats:
                            count += 1
                pct = count / substantive_count
                row.append(pct)
            else:
                row.append(0)
        rows.append(row)

    rows.append([''])

    # By Age Group
    add_section(rows, 'By Age Group - % of substantive respondents by each age group', False)
    header = ['Theme']
    age_groups = ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75+', 'Under 18', 'Not specified']
    for age in age_groups:
        age_answered = respondents_who_answered[respondents_who_answered['Age_Group'] == age]
        # Count substantive RESPONDENTS (not mentions)
        substantive_count = 0
        for val in age_answered.iloc[:, insight_col]:
            if pd.notna(val):
                cats = [c.strip() for c in str(val).split('|')]
                if any(c not in trivial_categories for c in cats):
                    substantive_count += 1
        if substantive_count > 0:
            header.append(f'{age}\n(n={substantive_count})')
    rows.append(header)

    for theme in all_themes:
        if pd.isna(theme):
            continue
        row = [theme]
        for age in age_groups:
            age_answered = respondents_who_answered[respondents_who_answered['Age_Group'] == age]
            # Count substantive respondents
            substantive_count = 0
            for val in age_answered.iloc[:, insight_col]:
                if pd.notna(val):
                    cats = [c.strip() for c in str(val).split('|')]
                    if any(c not in trivial_categories for c in cats):
                        substantive_count += 1
            if substantive_count > 0:
                # Count RESPONDENTS who have this theme
                count = 0
                for val in age_answered.iloc[:, insight_col]:
                    if pd.notna(val):
                        cats = [c.strip() for c in str(val).split('|')]
                        if theme in cats:
                            count += 1
                pct = count / substantive_count
                row.append(pct)
        rows.append(row)

    rows.append([''])

    # By NPS Tier
    add_section(rows, 'By NPS Tier - % of substantive respondents by each tier', False)
    header = ['Theme']
    nps_tiers = ['Promoter', 'Passive', 'Detractor']
    for nps in nps_tiers:
        nps_answered = respondents_who_answered[respondents_who_answered['NPS_Tier'] == nps]
        # Count substantive RESPONDENTS (not mentions)
        substantive_count = 0
        for val in nps_answered.iloc[:, insight_col]:
            if pd.notna(val):
                cats = [c.strip() for c in str(val).split('|')]
                if any(c not in trivial_categories for c in cats):
                    substantive_count += 1
        header.append(f'{nps}\n(n={substantive_count})')
    rows.append(header)

    for theme in all_themes:
        if pd.isna(theme):
            continue
        row = [theme]
        for nps in nps_tiers:
            nps_answered = respondents_who_answered[respondents_who_answered['NPS_Tier'] == nps]
            # Count substantive respondents
            substantive_count = 0
            for val in nps_answered.iloc[:, insight_col]:
                if pd.notna(val):
                    cats = [c.strip() for c in str(val).split('|')]
                    if any(c not in trivial_categories for c in cats):
                        substantive_count += 1
            if substantive_count > 0:
                # Count RESPONDENTS who have this theme
                count = 0
                for val in nps_answered.iloc[:, insight_col]:
                    if pd.notna(val):
                        cats = [c.strip() for c in str(val).split('|')]
                        if theme in cats:
                            count += 1
                pct = count / substantive_count
                row.append(pct)
            else:
                row.append(0)
        rows.append(row)

    # Add enhanced insights
    insights = generate_openended_insights(theme_counts, substantive_answers, total_answers)
    add_insights_section(rows, insights)

    return rows

def create_nps_sheet(q_num, question_text, score_col_idx):
    """Create sheet for NPS question (0-10 scale)"""
    rows = []

    # Title
    rows.append([f'Q#{q_num}: {question_text}'])
    rows.append([''])

    # PART 1: GENERAL ANALYSIS
    add_section(rows, 'GENERAL ANALYSIS - Overall Score Distribution', True)

    score_data = pd.to_numeric(responses.iloc[:, score_col_idx], errors='coerce')
    score_dist = score_data.value_counts().sort_index(ascending=False)
    avg_score = score_data.mean()
    total_responses = score_data.notna().sum()

    rows.append(['Score', 'Count', '% of Responses'])
    for score in range(10, -1, -1):  # 10 down to 0
        count = score_dist.get(score, 0)
        pct = count / total_responses if total_responses > 0 else 0
        rows.append([score, count, pct])

    rows.append([''])
    rows.append(['AVERAGE SCORE:', f'{avg_score:.2f} / 10.0'])
    rows.append([f'Total responses: {total_responses}'])

    # PART 2: DEMOGRAPHIC BREAKDOWNS
    add_section(rows, 'DEMOGRAPHIC ANALYSIS - Average Scores', True)

    # By Gender - ALL OPTIONS
    add_section(rows, 'By Gender', False)
    rows.append(['Gender', 'n', 'Average Score'])
    for gender in ['Female', 'Male', 'Other', 'Prefer not to say']:
        subset = responses[responses['Primary_Gender'] == gender]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([gender, n, avg])

    rows.append([''])

    # By Payer Segment
    add_section(rows, 'By Payer Segment', False)
    rows.append(['Payer Segment', 'n', 'Average Score'])
    for ps in ['Non_Payer', 'Low', 'Med', 'High', 'VIP']:
        subset = responses[responses['Payer_Segment'] == ps]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([ps, n, avg])

    rows.append([''])

    # By Age Group
    add_section(rows, 'By Age Group', False)
    rows.append(['Age Group', 'n', 'Average Score'])
    for age in ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75+', 'Under 18']:
        subset = responses[responses['Age_Group'] == age]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([age, n, avg])

    rows.append([''])

    # By NPS Tier
    add_section(rows, 'By NPS Tier', False)
    rows.append(['NPS Tier', 'n', 'Average Score'])
    for nps in ['Promoter', 'Passive', 'Detractor']:
        subset = responses[responses['NPS_Tier'] == nps]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([nps, n, avg])

    return rows

def create_rating_sheet(q_num, question_text, score_col_idx):
    """Create sheet for rating scale questions (1-5)"""
    rows = []

    # Title
    rows.append([f'Q#{q_num}: {question_text}'])
    rows.append([''])

    # PART 1: GENERAL ANALYSIS
    add_section(rows, 'GENERAL ANALYSIS - Overall Score Distribution', True)

    score_data = pd.to_numeric(responses.iloc[:, score_col_idx], errors='coerce')
    score_dist = score_data.value_counts().sort_index(ascending=False)
    avg_score = score_data.mean()
    total_responses = score_data.notna().sum()

    rows.append(['Score', 'Count', '% of Responses'])
    for score in range(5, 0, -1):
        count = score_dist.get(score, 0)
        pct = count / total_responses if total_responses > 0 else 0
        rows.append([score, count, pct])

    rows.append([''])
    rows.append(['AVERAGE SCORE:', f'{avg_score:.2f} / 5.0'])
    rows.append([f'Total responses: {total_responses}'])

    # PART 2: DEMOGRAPHIC BREAKDOWNS
    add_section(rows, 'DEMOGRAPHIC ANALYSIS - Average Scores', True)

    # By Gender - ALL OPTIONS
    add_section(rows, 'By Gender', False)
    rows.append(['Gender', 'n', 'Average Score'])
    for gender in ['Female', 'Male', 'Other', 'Prefer not to say']:
        subset = responses[responses['Primary_Gender'] == gender]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([gender, n, avg])

    rows.append([''])

    # By Payer Segment
    add_section(rows, 'By Payer Segment', False)
    rows.append(['Payer Segment', 'n', 'Average Score'])
    for ps in ['Non_Payer', 'Low', 'Med', 'High', 'VIP']:
        subset = responses[responses['Payer_Segment'] == ps]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([ps, n, avg])

    rows.append([''])

    # By Age Group
    add_section(rows, 'By Age Group', False)
    rows.append(['Age Group', 'n', 'Average Score'])
    for age in ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75+', 'Under 18']:
        subset = responses[responses['Age_Group'] == age]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([age, n, avg])

    rows.append([''])

    # By NPS Tier
    add_section(rows, 'By NPS Tier', False)
    rows.append(['NPS Tier', 'n', 'Average Score'])
    for nps in ['Promoter', 'Passive', 'Detractor']:
        subset = responses[responses['NPS_Tier'] == nps]
        subset_scores = pd.to_numeric(subset.iloc[:, score_col_idx], errors='coerce')
        n = subset_scores.notna().sum()
        if n > 0:
            avg = subset_scores.mean()
            rows.append([nps, n, avg])

    return rows

def create_multiselect_sheet(q_num, question_text, option_cols):
    """Create sheet for multi-select questions (Q8, Q9)"""
    rows = []

    # Title
    rows.append([f'Q#{q_num}: {question_text}'])
    rows.append([''])

    # PART 1: GENERAL ANALYSIS
    add_section(rows, 'GENERAL ANALYSIS - Selection Frequency', True)

    # Count respondents who selected at least one option
    respondents_who_answered = responses[
        responses.iloc[:, option_cols].notna().any(axis=1)
    ]
    total_who_answered = len(respondents_who_answered)

    # Count selections for each option (among those who answered)
    option_counts = []
    for col_idx in option_cols:
        option_name = metadata.iloc[col_idx]
        count = respondents_who_answered.iloc[:, col_idx].notna().sum()
        option_counts.append((option_name, count))

    # Sort by count
    option_counts.sort(key=lambda x: x[1], reverse=True)

    rows.append(['Option', 'Count', '% Selected'])
    for option, count in option_counts:
        # Store as decimal for Excel formatting
        pct = count / total_who_answered if total_who_answered > 0 else 0
        rows.append([option, count, pct])

    rows.append([''])
    rows.append([f'Total respondents who answered this question: {total_who_answered}'])
    rows.append([f'Total respondents in survey: {len(responses)}'])
    rows.append(['Note: Percentages show what % of respondents who answered this question selected each option'])

    # PART 2: DEMOGRAPHIC BREAKDOWNS
    add_section(rows, 'DEMOGRAPHIC ANALYSIS - All Options', True)

    # Use ALL options (not just top 10)
    all_options = option_counts

    # By Gender
    add_section(rows, 'By Gender - % who selected each option', False)
    header = ['Option']
    genders = ['Female', 'Male', 'Other', 'Prefer not to say', 'Multiple', 'Not specified']
    for g in genders:
        # Count only those of this gender who answered this question
        gender_answered = respondents_who_answered[respondents_who_answered['Primary_Gender'] == g]
        n = len(gender_answered)
        if n > 0:
            header.append(f'{g}\n(n={n})')
    rows.append(header)

    for option_name, _ in all_options:
        row = [option_name]
        # Find the column index for this option
        col_idx = None
        for idx in option_cols:
            if metadata.iloc[idx] == option_name:
                col_idx = idx
                break

        if col_idx is not None:
            for g in genders:
                gender_answered = respondents_who_answered[respondents_who_answered['Primary_Gender'] == g]
                n = len(gender_answered)
                if n > 0:
                    count = gender_answered.iloc[:, col_idx].notna().sum()
                    # Store as decimal for Excel formatting
                    pct = count / n
                    row.append(pct)
        rows.append(row)

    rows.append([''])

    # By Age Group
    add_section(rows, 'By Age Group - % who selected each option', False)
    header = ['Option']
    age_groups = ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75+', 'Under 18', 'Not specified']
    for age in age_groups:
        age_answered = respondents_who_answered[respondents_who_answered['Age_Group'] == age]
        n = len(age_answered)
        if n > 0:
            header.append(f'{age}\n(n={n})')
    rows.append(header)

    for option_name, _ in all_options:
        row = [option_name]
        col_idx = None
        for idx in option_cols:
            if metadata.iloc[idx] == option_name:
                col_idx = idx
                break

        if col_idx is not None:
            for age in age_groups:
                age_answered = respondents_who_answered[respondents_who_answered['Age_Group'] == age]
                n = len(age_answered)
                if n > 0:
                    count = age_answered.iloc[:, col_idx].notna().sum()
                    pct = count / n
                    row.append(pct)
        rows.append(row)

    rows.append([''])

    # By Payer Segment
    add_section(rows, 'By Payer Segment - % who selected each option', False)
    header = ['Option']
    payer_segs = ['Non_Payer', 'Low', 'Med', 'High', 'VIP']
    for ps in payer_segs:
        ps_answered = respondents_who_answered[respondents_who_answered['Payer_Segment'] == ps]
        n = len(ps_answered)
        header.append(f'{ps}\n(n={n})')
    rows.append(header)

    for option_name, _ in all_options:
        row = [option_name]
        col_idx = None
        for idx in option_cols:
            if metadata.iloc[idx] == option_name:
                col_idx = idx
                break

        if col_idx is not None:
            for ps in payer_segs:
                ps_answered = respondents_who_answered[respondents_who_answered['Payer_Segment'] == ps]
                n = len(ps_answered)
                count = ps_answered.iloc[:, col_idx].notna().sum() if n > 0 else 0
                pct = count / n if n > 0 else 0
                row.append(pct)
        rows.append(row)

    rows.append([''])

    # By NPS Tier
    add_section(rows, 'By NPS Tier - % who selected each option', False)
    header = ['Option']
    nps_tiers = ['Promoter', 'Passive', 'Detractor']
    for nps in nps_tiers:
        nps_answered = respondents_who_answered[respondents_who_answered['NPS_Tier'] == nps]
        n = len(nps_answered)
        header.append(f'{nps}\n(n={n})')
    rows.append(header)

    for option_name, _ in all_options:
        row = [option_name]
        col_idx = None
        for idx in option_cols:
            if metadata.iloc[idx] == option_name:
                col_idx = idx
                break

        if col_idx is not None:
            for nps in nps_tiers:
                nps_answered = respondents_who_answered[respondents_who_answered['NPS_Tier'] == nps]
                n = len(nps_answered)
                count = nps_answered.iloc[:, col_idx].notna().sum() if n > 0 else 0
                pct = count / n if n > 0 else 0
                row.append(pct)
        rows.append(row)

    # Add enhanced insights
    insights = generate_multiselect_insights(option_counts, respondents_who_answered, option_cols, metadata)
    add_insights_section(rows, insights)

    return rows

def create_categorical_sheet(q_num, question_text, data_col_idx):
    """Create sheet for categorical questions (Q10, Q11)"""
    rows = []

    # Title
    rows.append([f'Q#{q_num}: {question_text}'])
    rows.append([''])

    # PART 1: GENERAL ANALYSIS
    add_section(rows, 'GENERAL ANALYSIS - Response Distribution', True)

    cat_data = responses.iloc[:, data_col_idx]
    cat_counts = cat_data.value_counts()
    total_responses = cat_data.notna().sum()

    rows.append(['Response', 'Count', '% of Responses'])
    for response, count in cat_counts.items():
        if pd.notna(response):
            pct = count / total_responses
            rows.append([response, count, pct])

    rows.append([''])
    rows.append([f'Total responses: {total_responses}'])

    # PART 2: DEMOGRAPHIC BREAKDOWNS
    add_section(rows, 'DEMOGRAPHIC ANALYSIS', True)

    # By Gender
    add_section(rows, 'By Gender - % giving each response', False)
    header = ['Response']
    genders = ['Female', 'Male', 'Other', 'Prefer not to say', 'Multiple', 'Not specified']
    for g in genders:
        # Count only those of this gender who answered this question
        gender_subset = responses[responses['Primary_Gender'] == g]
        n = gender_subset.iloc[:, data_col_idx].notna().sum()
        if n > 0:
            header.append(f'{g}\n(n={n})')
    rows.append(header)

    for response in cat_counts.index:
        if pd.isna(response):
            continue
        row = [response]
        for g in genders:
            gender_subset = responses[responses['Primary_Gender'] == g]
            n = gender_subset.iloc[:, data_col_idx].notna().sum()
            if n > 0:
                count = (gender_subset.iloc[:, data_col_idx] == response).sum()
                pct = count / n  # Store as decimal for Excel formatting
                row.append(pct)
        rows.append(row)

    rows.append([''])

    # By Age Group
    add_section(rows, 'By Age Group - % giving each response', False)
    header = ['Response']
    age_groups = ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75+', 'Under 18', 'Not specified']
    for age in age_groups:
        age_subset = responses[responses['Age_Group'] == age]
        n = age_subset.iloc[:, data_col_idx].notna().sum()
        if n > 0:
            header.append(f'{age}\n(n={n})')
    rows.append(header)

    for response in cat_counts.index:
        if pd.isna(response):
            continue
        row = [response]
        for age in age_groups:
            age_subset = responses[responses['Age_Group'] == age]
            n = age_subset.iloc[:, data_col_idx].notna().sum()
            if n > 0:
                count = (age_subset.iloc[:, data_col_idx] == response).sum()
                pct = count / n
                row.append(pct)
        rows.append(row)

    rows.append([''])

    # By Payer Segment
    add_section(rows, 'By Payer Segment - % giving each response', False)
    header = ['Response']
    payer_segs = ['Non_Payer', 'Low', 'Med', 'High', 'VIP']
    for ps in payer_segs:
        ps_subset = responses[responses['Payer_Segment'] == ps]
        n = ps_subset.iloc[:, data_col_idx].notna().sum()
        header.append(f'{ps}\n(n={n})')
    rows.append(header)

    for response in cat_counts.index:
        if pd.isna(response):
            continue
        row = [response]
        for ps in payer_segs:
            ps_subset = responses[responses['Payer_Segment'] == ps]
            n = ps_subset.iloc[:, data_col_idx].notna().sum()
            count = (ps_subset.iloc[:, data_col_idx] == response).sum()
            pct = count / n if n > 0 else 0
            row.append(pct)
        rows.append(row)

    rows.append([''])

    # By NPS Tier
    add_section(rows, 'By NPS Tier - % giving each response', False)
    header = ['Response']
    nps_tiers = ['Promoter', 'Passive', 'Detractor']
    for nps in nps_tiers:
        nps_subset = responses[responses['NPS_Tier'] == nps]
        n = nps_subset.iloc[:, data_col_idx].notna().sum()
        header.append(f'{nps}\n(n={n})')
    rows.append(header)

    for response in cat_counts.index:
        if pd.isna(response):
            continue
        row = [response]
        for nps in nps_tiers:
            nps_subset = responses[responses['NPS_Tier'] == nps]
            n = nps_subset.iloc[:, data_col_idx].notna().sum()
            count = (nps_subset.iloc[:, data_col_idx] == response).sum()
            pct = count / n if n > 0 else 0
            row.append(pct)
        rows.append(row)

    # Add enhanced insights
    insights = generate_categorical_insights(cat_counts, total_responses)
    add_insights_section(rows, insights)

    return rows

def create_demographic_sheet(q_num, question_text, demo_type):
    """Create sheet for demographic questions (Q13 Age, Q14 Gender)"""
    rows = []

    # Title
    rows.append([f'Q#{q_num}: {question_text}'])
    rows.append([''])

    if demo_type == 'age':
        demo_col = 'Age_Group'
        demo_values = ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75+', 'Under 18', 'Not specified']
    else:  # gender
        demo_col = 'Primary_Gender'
        demo_values = ['Female', 'Male', 'Other', 'Prefer not to say', 'Not specified']

    # PART 1: GENERAL ANALYSIS
    add_section(rows, 'GENERAL ANALYSIS - Overall Distribution', True)

    demo_counts = responses[demo_col].value_counts()
    total_respondents = len(responses)  # Use total respondents so percentages add to 100%

    rows.append([demo_type.title(), 'Count', '% of Responses'])
    for val in demo_values:
        count = demo_counts.get(val, 0)
        pct = (count / total_respondents) if total_respondents > 0 else 0  # Store as decimal for Excel
        if count > 0:  # Only show if there are responses
            rows.append([val, count, pct])

    rows.append([''])
    rows.append([f'Total respondents: {total_respondents}'])

    # PART 2: DEMOGRAPHIC BREAKDOWNS
    add_section(rows, 'DEMOGRAPHIC ANALYSIS', True)

    if demo_type == 'age':
        # Age broken down by Gender and Payer Segment
        add_section(rows, 'By Gender - % of each gender in each age group', False)
        header = ['Age Group']
        genders = ['Female', 'Male', 'Other', 'Prefer not to say', 'Multiple', 'Not specified']
        for g in genders:
            n = (responses['Primary_Gender'] == g).sum()
            if n > 0:
                header.append(f'{g}\n(n={n})')
        rows.append(header)

        for age in demo_values:
            count = demo_counts.get(age, 0)
            if count > 0:
                row = [age]
                for g in genders:
                    n = (responses['Primary_Gender'] == g).sum()
                    if n > 0:
                        subset = responses[responses['Primary_Gender'] == g]
                        age_count = (subset['Age_Group'] == age).sum()
                        pct = age_count / n * 100
                        row.append(pct)
                rows.append(row)

        rows.append([''])

        add_section(rows, 'By Payer Segment - % of each segment in each age group', False)
        header = ['Age Group']
        payer_segs = ['Non_Payer', 'Low', 'Med', 'High', 'VIP']
        for ps in payer_segs:
            n = (responses['Payer_Segment'] == ps).sum()
            header.append(f'{ps}\n(n={n})')
        rows.append(header)

        for age in demo_values:
            count = demo_counts.get(age, 0)
            if count > 0:
                row = [age]
                for ps in payer_segs:
                    subset = responses[responses['Payer_Segment'] == ps]
                    n = len(subset)
                    age_count = (subset['Age_Group'] == age).sum()
                    pct = age_count / n * 100 if n > 0 else 0
                    row.append(pct)
                rows.append(row)

    else:  # gender
        # Gender broken down by Age and Payer Segment
        add_section(rows, 'By Age Group - % of each age group identifying as each gender', False)
        header = ['Gender']
        ages = ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74']
        for age in ages:
            n = (responses['Age_Group'] == age).sum()
            if n > 10:  # Only show if enough responses
                header.append(f'{age}\n(n={n})')
        rows.append(header)

        for gender in demo_values:
            count = demo_counts.get(gender, 0)
            if count > 0:
                row = [gender]
                for age in ages:
                    n = (responses['Age_Group'] == age).sum()
                    if n > 10:
                        subset = responses[responses['Age_Group'] == age]
                        gender_count = (subset['Primary_Gender'] == gender).sum()
                        pct = gender_count / n * 100
                        row.append(pct)
                rows.append(row)

        rows.append([''])

        add_section(rows, 'By Payer Segment - % of each segment identifying as each gender', False)
        header = ['Gender']
        payer_segs = ['Non_Payer', 'Low', 'Med', 'High', 'VIP']
        for ps in payer_segs:
            n = (responses['Payer_Segment'] == ps).sum()
            header.append(f'{ps}\n(n={n})')
        rows.append(header)

        for gender in demo_values:
            count = demo_counts.get(gender, 0)
            if count > 0:
                row = [gender]
                for ps in payer_segs:
                    subset = responses[responses['Payer_Segment'] == ps]
                    n = len(subset)
                    gender_count = (subset['Primary_Gender'] == gender).sum()
                    pct = gender_count / n * 100 if n > 0 else 0
                    row.append(pct)
                rows.append(row)

    return rows

# ============================================================
# ENHANCED INSIGHTS HELPERS
# ============================================================

def generate_demographic_insights(category_name, overall_pct, segment_data, segment_type):
    """
    Generate insights comparing demographic segments to overall population

    Args:
        category_name: Name of the category/option being analyzed
        overall_pct: Overall percentage for this category (as decimal, e.g., 0.386)
        segment_data: Dict mapping segment name to (count, total, percentage) tuples
        segment_type: Type of segment ('gender', 'payer', 'age', 'nps')

    Returns:
        List of insight strings
    """
    insights = []

    # Filter segments with sufficient data (n >= 30)
    valid_segments = {seg: data for seg, data in segment_data.items()
                     if data[1] >= 30}  # data[1] is total count

    if not valid_segments:
        return ["Sample sizes too small for reliable insights (all segments <30 respondents)"]

    # Calculate deviations from overall
    deviations = []
    for seg, (count, total, pct) in valid_segments.items():
        deviation = pct - overall_pct
        if abs(deviation) >= 0.10:  # 10% or more deviation
            deviations.append((seg, deviation, pct))

    # Sort by absolute deviation
    deviations.sort(key=lambda x: abs(x[1]), reverse=True)

    # Generate insights
    if deviations:
        # Top positive deviations
        positive = [d for d in deviations if d[1] > 0]
        if positive:
            seg, dev, pct = positive[0]
            insights.append(f"**{seg}** shows highest affinity: {pct:.0%} vs {overall_pct:.0%} overall (+{dev:.0%})")

        # Top negative deviations
        negative = [d for d in deviations if d[1] < 0]
        if negative:
            seg, dev, pct = negative[-1]  # Most negative
            insights.append(f"**{seg}** shows lowest affinity: {pct:.0%} vs {overall_pct:.0%} overall ({dev:.0%})")

    # Note small sample sizes
    small_samples = [seg for seg, data in segment_data.items() if data[1] < 30]
    if small_samples:
        insights.append(f"⚠ Small sample sizes for: {', '.join(small_samples)} - interpret with caution")

    return insights if insights else ["No significant deviations from overall population"]


def add_insights_section(rows, insight_texts):
    """Add an Enhanced Insights section to the output"""
    rows.append([''])
    rows.append(['═' * 100])
    rows.append(['ENHANCED INSIGHTS'])
    rows.append(['═' * 100])
    for insight in insight_texts:
        rows.append([insight])
    rows.append([''])


def generate_multiselect_insights(option_counts, respondents_who_answered, option_cols, metadata):
    """Generate insights for multiselect questions"""
    insights = []

    # Category dominance
    if option_counts:
        top_option, top_count = option_counts[0]
        bottom_option, bottom_count = option_counts[-1]
        total = len(respondents_who_answered)

        insights.append(f"📊 **Category Dominance:**")
        insights.append(f"   • Most popular: '{top_option}' - {top_count}/{total} ({top_count/total:.0%}) of respondents")
        insights.append(f"   • Least popular: '{bottom_option}' - {bottom_count}/{total} ({bottom_count/total:.0%}) of respondents")
        insights.append("")

    # Demographic sample sizes
    insights.append("📈 **Sample Size Notes:**")

    # Check gender segments
    genders = ['Female', 'Male', 'Other', 'Prefer not to say', 'Multiple', 'Not specified']
    small_genders = []
    for g in genders:
        gender_count = len(respondents_who_answered[respondents_who_answered['Primary_Gender'] == g])
        if 0 < gender_count < 30:
            small_genders.append(f"{g} (n={gender_count})")

    if small_genders:
        insights.append(f"   ⚠ Small gender segments: {', '.join(small_genders)}")
    else:
        insights.append(f"   ✓ All gender segments have sufficient sample sizes (n≥30)")

    # Check age segments
    age_groups = ['18-24', '25-34', '35-44', '45-54', '55-64', '65-74', '75+', 'Under 18', 'Not specified']
    small_ages = []
    for age in age_groups:
        age_count = len(respondents_who_answered[respondents_who_answered['Age_Group'] == age])
        if 0 < age_count < 30:
            small_ages.append(f"{age} (n={age_count})")

    if small_ages:
        insights.append(f"   ⚠ Small age segments: {', '.join(small_ages)}")
    else:
        insights.append(f"   ✓ All age segments have sufficient sample sizes (n≥30)")

    insights.append("")
    insights.append("💡 **Analysis Tip:** Look for demographic segments that deviate ±10% or more from the overall percentage to identify unique preferences.")

    return insights


def generate_openended_insights(theme_counts, substantive_answers, total_answers):
    """Generate insights for open-ended questions"""
    insights = []

    # Category dominance
    if len(theme_counts) > 0:
        top_themes = theme_counts.head(3)
        insights.append(f"📊 **Top Themes:**")
        for theme, count in top_themes.items():
            pct = count / substantive_answers if substantive_answers > 0 else 0
            insights.append(f"   • '{theme}' - {count} mentions ({pct:.0%} of substantive responses)")
        insights.append("")

    # Response rate
    response_rate = total_answers / len(responses) if len(responses) > 0 else 0
    substantive_rate = substantive_answers / total_answers if total_answers > 0 else 0

    insights.append(f"📈 **Response Quality:**")
    insights.append(f"   • Response rate: {response_rate:.0%} of respondents provided answers")
    insights.append(f"   • Substantive rate: {substantive_rate:.0%} of answers contained meaningful feedback")
    insights.append("")

    insights.append("💡 **Multi-Category Note:** Percentages may sum to >100% because respondents can have multiple categories.")

    return insights


def generate_categorical_insights(value_counts, total_responses):
    """Generate insights for categorical questions"""
    insights = []

    # Category distribution
    if len(value_counts) > 0:
        top_category, top_count = list(value_counts.items())[0]
        insights.append(f"📊 **Distribution:**")
        insights.append(f"   • Most common: '{top_category}' - {top_count}/{total_responses} ({top_count/total_responses:.0%})")

        if len(value_counts) > 1:
            bottom_category, bottom_count = list(value_counts.items())[-1]
            insights.append(f"   • Least common: '{bottom_category}' - {bottom_count}/{total_responses} ({bottom_count/total_responses:.0%})")
        insights.append("")

    insights.append("💡 **Analysis Tip:** Compare demographic segments to identify groups with distinct preferences.")

    return insights


# ============================================================
# CHART CREATION HELPERS
# ============================================================

def add_chart_for_table(ws, start_row, start_col, num_data_rows, num_data_cols, chart_title, chart_position, use_stacked=False, skip_count_column=False):
    """
    Add a bar chart for a data table

    Args:
        ws: worksheet
        start_row: row number where data starts (1-indexed, after header)
        start_col: column number where labels are (1-indexed, typically column 1)
        num_data_rows: number of data rows
        num_data_cols: number of data columns to chart (excluding label column)
        chart_title: title for the chart
        chart_position: cell position for chart (e.g., 'A20')
        use_stacked: DEPRECATED - all charts now use clustered/grouped bars
        skip_count_column: If True, skip the count column (col 2) and only chart percentage columns
    """
    if num_data_rows == 0 or num_data_cols == 0:
        return

    chart = BarChart()
    chart.title = chart_title
    chart.style = 10

    # Vary colors by point for better visualization (when single series)
    # if num_data_cols == 1:
    chart.varyColors = True

    # Determine chart type based on number of categories
    if num_data_rows <= 5:
        chart.type = "col"  # Vertical bar chart for ≤5 categories
        chart.grouping = "clustered"  # Never stack - always use grouped bars
    else:
        # Horizontal bar for >5 categories (better for long labels)
        chart.type = "bar"
        chart.direction = "bar"
        chart.grouping = "clustered"  # Never stack - always use grouped bars

    # --- FIX: Exclude count (n) columns even if table layout changes ---
    # Determine which columns contain numeric averages or percentages
    valid_cols = []
    header_row_idx = start_row - 1

    for col in range(start_col + 1, start_col + num_data_cols + 2):
        header_value = ws.cell(header_row_idx, col).value
        if header_value is None:
            continue
        # Skip columns containing any form of n-count
        if isinstance(header_value, str) and (
            header_value.strip().lower() == "n" or "(n=" in header_value.lower()
        ):
            continue
        valid_cols.append(col)

    # If no columns remain, do nothing
    if not valid_cols:
        return


    # Data reference (values)
    # If skip_count_column, start from column 3 (percentage) instead of column 2 (count)
    data_start_col = start_col + 2 if skip_count_column else start_col + 1
    # data = Reference(ws, min_col=data_start_col, max_col=data_start_col + num_data_cols - 1,
    #                  min_row=start_row-1, max_row=start_row+num_data_rows-1)
    
    data = Reference(ws,
                 min_col=min(valid_cols),
                 max_col=max(valid_cols),
                 min_row=start_row-1,
                 max_row=start_row+num_data_rows-1)


    # Category reference (labels) - always from start_col
    cats = Reference(ws, min_col=start_col, min_row=start_row, max_row=start_row+num_data_rows-1)

    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    # Enable data labels showing percentages
    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True

    # Set axis limits - for multi-category questions, values can exceed 100%
    # Set max to 150% (1.5) or the max value, whichever is greater
    # Automatically compute the max value in the chart data
    values = []
    for col in range(data_start_col, data_start_col + num_data_cols):
        for row in range(start_row, start_row + num_data_rows):
            cell_value = ws.cell(row, col).value
            if isinstance(cell_value, (int, float)):
                values.append(cell_value)

    max_val = max(values) if values else 1.0

    # Always start at 0
    chart.y_axis.scaling.min = 0.0

    # If max is <= 1.0 (i.e., percentages), set to 1.0 (100%)
    # Otherwise, add a little padding
    if max_val <= 1:
        chart.y_axis.scaling.max = 1.0    # 100%
    else:
        chart.y_axis.scaling.max = max_val * 1.1  # +10% padding


    # Set size
    chart.width = 15
    chart.height = 10

    ws.add_chart(chart, chart_position)


def find_table_by_header(ws, header_text):
    """
    Find a table by searching for its header text
    Returns (header_row, num_cols, data_rows) or None if not found
    """
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        cell = row[0]
        if cell.value and isinstance(cell.value, str) and header_text in cell.value:
            # Found the section header, now find the actual data table header
            # Skip to next non-empty row (the actual table header)
            table_header_row = None
            for search_row in range(row_idx + 1, min(row_idx + 5, ws.max_row + 1)):
                first_cell = ws.cell(search_row, 1).value
                if first_cell and not str(first_cell).startswith('═'):
                    table_header_row = search_row
                    break

            if table_header_row:
                # Count number of columns
                num_cols = 0
                for col_idx in range(1, 20):  # Check up to 20 columns
                    if ws.cell(table_header_row, col_idx).value:
                        num_cols = col_idx
                    else:
                        break

                # Count data rows (stop at empty row or section divider)
                num_data_rows = 0
                for data_row in range(table_header_row + 1, ws.max_row + 1):
                    first_cell = ws.cell(data_row, 1).value
                    if not first_cell or str(first_cell).strip() == '' or str(first_cell).startswith('═'):
                        break
                    num_data_rows += 1

                return (table_header_row, num_cols, num_data_rows)

    return None


def add_charts_to_sheet(ws, q_num, question_type, question_text):
    """
    Add charts to a question sheet

    Args:
        ws: worksheet
        q_num: question number (for title)
        question_type: 'multiselect', 'categorical', 'open-ended', 'rating', 'nps', or 'demographic'
        question_text: actual question text for chart title
    """
    # Extract short version of question for title (first 50 chars)
    short_question = question_text[:50] + "..." if len(question_text) > 50 else question_text

    # Add chart for GENERAL ANALYSIS (overall distribution)
    general_table = None
    if question_type == 'multiselect':
        general_table = find_table_by_header(ws, "GENERAL ANALYSIS - Selection Frequency")
    elif question_type == 'categorical':
        general_table = find_table_by_header(ws, "GENERAL ANALYSIS - Response Distribution")
    elif question_type == 'open-ended':
        general_table = find_table_by_header(ws, "GENERAL ANALYSIS - Overall Theme Distribution")
    elif question_type in ['rating', 'nps']:
        general_table = find_table_by_header(ws, "GENERAL ANALYSIS")
    elif question_type == 'demographic':
        general_table = find_table_by_header(ws, "GENERAL ANALYSIS - Distribution")

    if general_table:
        header_row, num_cols, num_data_rows = general_table
        # For general analysis, show ONLY the percentage column(s), skip count
        if num_data_rows > 0 and num_cols >= 3:
            chart_row = header_row + num_data_rows + 2
            chart_position = f'A{chart_row}'
            # For most tables: col1=label, col2=count, col3=percentage (or more % columns)
            # We skip the count column and chart only percentage columns

            # For open-ended: [Theme, Mentions, % of All, % of Substantive] - chart col 3 & 4
            # For multiselect: [Option, Count, % Selected] - chart col 3
            # For categorical: [Response, Count, % of Responses] - chart col 3

            # Number of percentage columns = total columns - label col - count col
            num_pct_cols = max(1, num_cols - 2)  # At least 1 percentage column

            add_chart_for_table(ws, header_row + 1, 1, num_data_rows, num_pct_cols,
                              f"Q{q_num}: {short_question}", chart_position,
                              use_stacked=False, skip_count_column=True)

    # For demographic questions, also add charts for cross-tabulations
    if question_type == 'demographic':
        # Add chart for By Payer Segment
        payer_table = find_table_by_header(ws, "By Payer Segment")
        if payer_table:
            header_row, num_cols, num_data_rows = payer_table
            if num_data_rows > 0 and num_cols > 1:
                chart_row = header_row + num_data_rows + 2
                chart_position = f'A{chart_row}'
                add_chart_for_table(ws, header_row + 1, 1, num_data_rows, num_cols - 1,
                                  f"Q{q_num}: {short_question} - By Payer Segment", chart_position, use_stacked=True)

        # Add chart for By Age Group
        age_table = find_table_by_header(ws, "By Age Group")
        if age_table:
            header_row, num_cols, num_data_rows = age_table
            if num_data_rows > 0 and num_cols > 1:
                chart_row = header_row + num_data_rows + 2
                chart_position = f'A{chart_row}'
                add_chart_for_table(ws, header_row + 1, 1, num_data_rows, num_cols - 1,
                                  f"Q{q_num}: {short_question} - By Age Group", chart_position, use_stacked=True)
        return  # Don't add demographic breakdown charts for demographic questions

    # Add demographic breakdown charts for all other question types
    # Find and add chart for Gender breakdown
    gender_table = find_table_by_header(ws, "By Gender")
    if gender_table:
        header_row, num_cols, num_data_rows = gender_table
        if num_data_rows > 0 and num_cols > 1:
            chart_row = header_row + num_data_rows + 2
            chart_position = f'A{chart_row}'
            add_chart_for_table(ws, header_row + 1, 1, num_data_rows, num_cols - 1,
                              f"Q{q_num}: {short_question} - By Gender", chart_position, use_stacked=True)

    # Find and add chart for Payer Segment breakdown
    payer_table = find_table_by_header(ws, "By Payer Segment")
    if payer_table:
        header_row, num_cols, num_data_rows = payer_table
        if num_data_rows > 0 and num_cols > 1:
            chart_row = header_row + num_data_rows + 2
            chart_position = f'A{chart_row}'
            add_chart_for_table(ws, header_row + 1, 1, num_data_rows, num_cols - 1,
                              f"Q{q_num}: {short_question} - By Payer Segment", chart_position, use_stacked=True)

    # Find and add chart for Age Group breakdown
    age_table = find_table_by_header(ws, "By Age Group")
    if age_table:
        header_row, num_cols, num_data_rows = age_table
        if num_data_rows > 0 and num_cols > 1:
            chart_row = header_row + num_data_rows + 2
            chart_position = f'A{chart_row}'
            add_chart_for_table(ws, header_row + 1, 1, num_data_rows, num_cols - 1,
                              f"Q{q_num}: {short_question} - By Age Group", chart_position, use_stacked=True)

    # Find and add chart for NPS Tier breakdown
    nps_table = find_table_by_header(ws, "By NPS Tier")
    if nps_table:
        header_row, num_cols, num_data_rows = nps_table
        if num_data_rows > 0 and num_cols > 1:
            chart_row = header_row + num_data_rows + 2
            chart_position = f'A{chart_row}'
            add_chart_for_table(ws, header_row + 1, 1, num_data_rows, num_cols - 1,
                              f"Q{q_num}: {short_question} - By NPS Tier", chart_position, use_stacked=True)

# ============================================================
# CREATE EXCEL FILE
# ============================================================
writer = pd.ExcelWriter(output_file, engine='openpyxl')

print("Creating sheets...")

# SUMMARY: Respondent Counts
print("  ✓ SUMMARY - Respondent Counts")
summary_rows = create_respondent_summary_sheet()
pd.DataFrame(summary_rows).to_excel(writer, sheet_name='SUMMARY - Respondent Counts', index=False, header=False)

# Q2: NPS Score
print("  ✓ Q02 - NPS Score")
q2_rows = create_nps_sheet(2, "How satisfied are you with Match Masters overall? (0-10)", 3)
pd.DataFrame(q2_rows).to_excel(writer, sheet_name='Q02 - NPS Score', index=False, header=False)

# Q3: What did you like?
print("  ✓ Q03 - What did you like?")
q3_rows = create_openended_sheet(3, "What did you like most about your experience?", 5, 6)
pd.DataFrame(q3_rows).to_excel(writer, sheet_name='Q03 - What did you like', index=False, header=False)

# Q4: How can we improve?
print("  ✓ Q04 - How can we improve?")
q4_rows = create_openended_sheet(4, "How can we improve your experience?", 7, 8)
pd.DataFrame(q4_rows).to_excel(writer, sheet_name='Q04 - How can we improve', index=False, header=False)

# Q5: Technical Satisfaction
print("  ✓ Q05 - Technical Satisfaction")
q5_rows = create_rating_sheet(5, "How satisfied are you with technical performance?", 9)
pd.DataFrame(q5_rows).to_excel(writer, sheet_name='Q05 - Technical Satisfaction', index=False, header=False)

# Q6: Fairness Satisfaction
print("  ✓ Q06 - Fairness Satisfaction")
q6_rows = create_rating_sheet(6, "How satisfied are you with fairness of competition?", 11)
pd.DataFrame(q6_rows).to_excel(writer, sheet_name='Q06 - Fairness Satisfaction', index=False, header=False)

# Q7: Progression Satisfaction
print("  ✓ Q07 - Progression Satisfaction")
q7_rows = create_rating_sheet(7, "How satisfied are you with progression and league climbing?", 13)
pd.DataFrame(q7_rows).to_excel(writer, sheet_name='Q07 - Progression Satisfaction', index=False, header=False)

# Q8: Play Motivations (multi-select)
print("  ✓ Q08 - Play Motivations")
q8_cols = list(range(15, 29))  # Columns 15-28
q8_rows = create_multiselect_sheet(8, "Which of these make you want to keep playing the game regularly?", q8_cols)
pd.DataFrame(q8_rows).to_excel(writer, sheet_name='Q08 - Play Motivations', index=False, header=False)

# Q9: Favorite Modes (multi-select)
print("  ✓ Q09 - Favorite Modes")
q9_cols = list(range(29, 39))  # Columns 29-38
q9_rows = create_multiselect_sheet(9, "Which of these daily modes do you enjoy the most?", q9_cols)
pd.DataFrame(q9_rows).to_excel(writer, sheet_name='Q09 - Favorite Modes', index=False, header=False)

# Q10: Game Mode Variety
print("  ✓ Q10 - Game Mode Variety")
q10_rows = create_categorical_sheet(10, "How do you feel about the amount of different game modes available?", 39)
pd.DataFrame(q10_rows).to_excel(writer, sheet_name='Q10 - Game Mode Variety', index=False, header=False)

# Q11: Update Frequency
print("  ✓ Q11 - Update Frequency")
q11_rows = create_categorical_sheet(11, "How do you feel about the frequency of new things released?", 40)
pd.DataFrame(q11_rows).to_excel(writer, sheet_name='Q11 - Update Frequency', index=False, header=False)

# Q12: One thing to change
print("  ✓ Q12 - One thing to change")
q12_rows = create_openended_sheet(12, "If exactly ONE thing could be changed or added to the game what would it be?", 41, 42)
pd.DataFrame(q12_rows).to_excel(writer, sheet_name='Q12 - One thing to change', index=False, header=False)

# Q13: Age Distribution
print("  ✓ Q13 - Age Distribution")
q13_rows = create_demographic_sheet(13, "How old are you?", 'age')
pd.DataFrame(q13_rows).to_excel(writer, sheet_name='Q13 - Age Distribution', index=False, header=False)

# Q14: Gender Distribution
print("  ✓ Q14 - Gender Distribution")
q14_rows = create_demographic_sheet(14, "What gender do you identify as?", 'gender')
pd.DataFrame(q14_rows).to_excel(writer, sheet_name='Q14 - Gender Distribution', index=False, header=False)

# Q15: Final comments
print("  ✓ Q15 - Final comments")
q15_rows = create_openended_sheet(15, "Anything else you'd like to share before we wrap up?", 48, 49)
pd.DataFrame(q15_rows).to_excel(writer, sheet_name='Q15 - Final comments', index=False, header=False)

# Save
writer.close()

# Apply formatting
print("\nApplying Excel formatting...")
wb = load_workbook(output_file)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]

    # Determine column types and adjust widths
    for col_idx, column in enumerate(ws.columns, 1):
        column = list(column)
        col_letter = get_column_letter(col_idx)

        # Build a map of which rows contain which type of data
        # Scan all rows to find section headers
        section_map = {}  # Maps row_idx -> 'percentage' or 'average_score'
        current_section = None

        for check_idx in range(len(column)):
            cell_val = column[check_idx].value
            if cell_val and isinstance(cell_val, str):
                # Check for average score section headers
                if 'Average Score' in cell_val:
                    current_section = 'average_score'
                # Check for percentage section headers
                elif any(pattern in cell_val for pattern in ['% of', '% who', '% giving', '% Selected', '% Responses', '% Substantive']):
                    current_section = 'percentage'
                # If we hit a new major section, reset
                elif cell_val.startswith('═'):  # Section divider
                    current_section = None
            section_map[check_idx] = current_section

        # Calculate max width and apply formatting
        max_length = 0
        for cell_idx, cell in enumerate(column):
            try:
                cell_value = cell.value
                original_value = cell_value  # Keep for width calculation
                section_type = section_map.get(cell_idx)

                if cell_value is not None and isinstance(cell_value, (int, float)):
                    # Apply formatting based on section type
                    if section_type == 'percentage':
                        # Percentage section: already stored as decimal (0.602 for 60.2%)
                        # Just apply percentage format without dividing
                        if -0.01 <= cell_value <= 2.0:  # Valid percentage range (0-200%)
                            cell.number_format = '0.0%'
                    elif section_type == 'average_score':
                        # Average score section: keep as-is, format with 2 decimals
                        if 0 <= cell_value <= 10:  # Valid score range
                            cell.number_format = '0.00'

                # Calculate length for width (use original value before conversion)
                if isinstance(original_value, (int, float)):
                    if section_type == 'percentage':
                        value_str = f"{original_value * 100:.1f}%"  # e.g., "38.8%" or "0.7%"
                    else:
                        value_str = f"{original_value:.2f}"
                else:
                    value_str = str(original_value) if original_value else ""

                if len(value_str) > max_length:
                    max_length = len(value_str)
            except:
                pass

        # Set column width (minimum 10, maximum 60)
        adjusted_width = max(min(max_length + 3, 60), 10)
        ws.column_dimensions[col_letter].width = adjusted_width

# Add charts to all question sheets
print("\nAdding charts to sheets...")
question_sheets = {
    'Q02 - NPS Score': ('2', 'nps', 'How satisfied are you with Match Masters overall?'),
    'Q03 - What did you like': ('3', 'open-ended', 'What did you like most about your experience?'),
    'Q04 - How can we improve': ('4', 'open-ended', 'How can we improve your experience?'),
    'Q05 - Technical Satisfaction': ('5', 'rating', 'How satisfied are you with technical performance?'),
    'Q06 - Fairness Satisfaction': ('6', 'rating', 'How satisfied are you with fairness of competition?'),
    'Q07 - Progression Satisfaction': ('7', 'rating', 'How satisfied are you with progression and league climbing?'),
    'Q08 - Play Motivations': ('8', 'multiselect', 'Which of these make you want to keep playing?'),
    'Q09 - Favorite Modes': ('9', 'multiselect', 'What are your favorite game modes?'),
    'Q10 - Game Mode Variety': ('10', 'categorical', 'Is there enough variety in game modes?'),
    'Q11 - Update Frequency': ('11', 'categorical', 'How do you feel about update frequency?'),
    'Q12 - One thing to change': ('12', 'open-ended', 'If you could change one thing, what would it be?'),
    'Q13 - Age Distribution': ('13', 'demographic', 'What is your age group?'),
    'Q14 - Gender Distribution': ('14', 'demographic', 'What is your gender?'),
    'Q15 - Final comments': ('15', 'open-ended', 'Any final comments or suggestions?'),
}

for sheet_name, (q_num, q_type, q_text) in question_sheets.items():
    if sheet_name in wb.sheetnames:
        try:
            ws = wb[sheet_name]
            add_charts_to_sheet(ws, q_num, q_type, q_text)
            print(f"  ✓ Added charts to {sheet_name}")
        except Exception as e:
            print(f"  ⚠ Could not add charts to {sheet_name}: {e}")

wb.save(output_file)

print("\n" + "="*80)
print("MULTI-CATEGORY ANALYSIS CREATED!")
print("="*80)
print(f"\nOutput file: {output_file}")
print(f"Sheets created: {len(wb.sheetnames)}")
print("\nAll questions included:")
print("  ✓ Q02 - NPS Score (0-10 rating)")
print("  ✓ Q03 - What did you like? (MULTI-CATEGORY)")
print("  ✓ Q04 - How can we improve? (MULTI-CATEGORY)")
print("  ✓ Q05 - Technical Satisfaction")
print("  ✓ Q06 - Fairness Satisfaction")
print("  ✓ Q07 - Progression Satisfaction")
print("  ✓ Q08 - Play Motivations (multi-select)")
print("  ✓ Q09 - Favorite Modes (multi-select)")
print("  ✓ Q10 - Game Mode Variety")
print("  ✓ Q11 - Update Frequency")
print("  ✓ Q12 - One thing to change (MULTI-CATEGORY)")
print("  ✓ Q13 - Age Distribution")
print("  ✓ Q14 - Gender Distribution (Female, Male, Other, Prefer not to say)")
print("  ✓ Q15 - Final comments (MULTI-CATEGORY)")
print("\nKey features:")
print("  ✓ MULTI-CATEGORY support - responses can have 2-3 categories")
print("  ✓ Each category mention counted separately")
print("  ✓ Example: 'Graphics | Rewards' counts as 2 mentions")
print("  ✓ Uses REAL payer_segment data from full_data.xlsx")
print("  ✓ One sheet per question")
print("  ✓ General analysis first (overall counts/percentages)")
print("  ✓ Then demographic breakdowns (Gender, Payer Segment, NPS)")
print("  ✓ Gender includes ALL options from Q14")
print("\n✅ Ready to use!")

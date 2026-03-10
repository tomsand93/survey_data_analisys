from openpyxl import load_workbook
from pathlib import Path
import os

base_dir   = Path(os.path.dirname(os.path.abspath(__file__))).parent
output_dir = base_dir / "output"

INPUT_XLSX  = output_dir / "MULTI_CATEGORY_ANALYSIS.xlsx"
OUTPUT_XLSX = output_dir / "MULTI_CATEGORY_ANALYSIS_zeros.xlsx"

wb = load_workbook(INPUT_XLSX)

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, (int, float)) and cell.value == 0:
                cell.value = None

wb.save(OUTPUT_XLSX)
